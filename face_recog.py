import face_recognition
import cv2
from openpyxl import Workbook
import datetime
import timeloop
import time


# Get a reference to webcam #0 (the default one)
video_capture = cv2.VideoCapture(0)

# Create a woorksheet
book = Workbook()
sheet = book.active

# Load images.

image_1 = face_recognition.load_image_file("1.jpg")
image_1_face_encoding = face_recognition.face_encodings(image_1)[0]

image_5 = face_recognition.load_image_file("5.jpg")
image_5_face_encoding = face_recognition.face_encodings(image_5)[0]

image_7 = face_recognition.load_image_file("7.jpg")
image_7_face_encoding = face_recognition.face_encodings(image_7)[0]

image_3 = face_recognition.load_image_file("3.jpg")
image_3_face_encoding = face_recognition.face_encodings(image_3)[0]

image_4 = face_recognition.load_image_file("4.jpg")
image_4_face_encoding = face_recognition.face_encodings(image_4)[0]

image_8 = face_recognition.load_image_file("8.jpg")
image_8_face_encoding = face_recognition.face_encodings(image_8)[0]

image_9 = face_recognition.load_image_file("9.jpg")
image_9_face_encoding = face_recognition.face_encodings(image_9)[0]

image_10 = face_recognition.load_image_file("10.jpg")
image_10_face_encoding = face_recognition.face_encodings(image_10)[0]

image_11 = face_recognition.load_image_file("11.jpg")
image_11_face_encoding = face_recognition.face_encodings(image_11)[0]

image_12 = face_recognition.load_image_file("12.jpg")
image_12_face_encoding = face_recognition.face_encodings(image_12)[0]

image_13 = face_recognition.load_image_file("13.jpg")
image_13_face_encoding = face_recognition.face_encodings(image_13)[0]

image_14 = face_recognition.load_image_file("14.jpg")
image_14_face_encoding = face_recognition.face_encodings(image_14)[0]

image_15 = face_recognition.load_image_file("15.jpg")
image_15_face_encoding = face_recognition.face_encodings(image_15)[0]

image_16 = face_recognition.load_image_file("16.jpg")
image_16_face_encoding = face_recognition.face_encodings(image_16)[0]

# Create arrays of known face encodings and their names
known_face_encodings = [

    image_1_face_encoding,
    image_5_face_encoding,
    image_7_face_encoding,
    image_3_face_encoding,
    image_4_face_encoding,
    image_8_face_encoding,
    image_9_face_encoding,
    image_10_face_encoding,
    image_11_face_encoding,
    image_12_face_encoding,
    image_13_face_encoding,
    image_14_face_encoding,
    image_15_face_encoding,
    image_16_face_encoding

]
known_face_names = [

    "1",
    "5",
    "7",
    "3",
    "4",
    "8",
    "9",
    "10",
    "11",
    "12",
    "13",
    "14",
    "15",
    "16"
]

# Initialize some variables
face_locations = []
face_encodings = []
face_names = []
process_this_frame = True

# Load present date and time
now = datetime.datetime.now()
today = now.day
month = now.month

while True:
    # Grab a single frame of video
    ret, frame = video_capture.read()

    #  print ("frame : ", frame)
    #  print("type of frame : ", type(frame))

    # Resize frame of video to 1/4 size for faster face recognition processing
    try:
        #  small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)
        small_frame = cv2.resize(frame, (250, 250), fx=0.25, fy=0.25)
    except Exception as e:
        print(str(e))

    # Convert the image from BGR color (which OpenCV uses) to RGB color (which face_recognition uses)
    rgb_small_frame = small_frame[:, :, ::-1]

    # Only process every other frame of video to save time
    if process_this_frame:
        # Find all the faces and face encodings in the current frame of video
        face_locations = face_recognition.face_locations(rgb_small_frame)
        face_encodings = face_recognition.face_encodings(rgb_small_frame, face_locations)

    face_names = []
    name = "Unknown"
    for face_encoding in face_encodings:
        # See if the face is a match for the known face(s)
        matches = face_recognition.compare_faces(known_face_encodings, face_encoding)
        name = "Unknown"

        # If a match was found in known_face_encodings, just use the first one.
        if True in matches:
            first_match_index = matches.index(True)
            name = known_face_names[first_match_index]
            # Assign attendance
            if int(name) in range(1, 61):
                    sheet.cell(row=int(name), column=int(today)).value = "Present"

            else:
                sheet.cell(row=int(name), column=int(today)).value = "Absent"

    face_names.append(name)

    process_this_frame = not process_this_frame

    # Display the results
    (top, right, bottom, left) = (0, 0, 0, 0)
    for (top, right, bottom, left), name in zip(face_locations, face_names):
        # Scale back up face locations since the frame we detected in was scaled to 1/4 size
        top *= 4
        right *= 4
        bottom *= 4
        left *= 4

    # Draw a box around the face
    cv2.rectangle(frame, (left, top), (right, bottom), (0, 0, 255), 2)

    # Draw a label with a name below the face
    cv2.rectangle(frame, (left, bottom - 35), (right, bottom), (0, 0, 255), cv2.FILLED)
    font = cv2.FONT_HERSHEY_DUPLEX
    cv2.putText(frame, name, (left + 6, bottom - 6), font, 1.0, (255, 255, 255), 1)

    # Display the resulting image
    try:
        cv2.imshow('frame', frame)
    except Exception as e:
        print(str(e))

    # Save Woorksheet as present month
    book.save(str(month) + '.xlsx')

    # Hit 'q' on the keyboard to quit!
    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

# Release handle to the webcam
video_capture.release()
cv2.destroyAllWindows()
# notification
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders

email_user = 'saurabhchordiya672@gmail.com'
email_password = '8208425554'
email_send = ['shrinivasgenge1403@gmail.com','gengeshri90@gmail.com']


subject = 'python'

msg = MIMEMultipart()
msg['From'] = email_user
msg['To'] = ",".join(email_send)
msg['Subject'] = subject

body = 'Hi there, sending this email from Python!'
msg.attach(MIMEText(body,'plain'))

filename='3.xlsx'
attachment  =open(filename,'rb')

part = MIMEBase('application','octet-stream')
part.set_payload((attachment).read())
encoders.encode_base64(part)
part.add_header('Content-Disposition',"attachment; filename= "+filename)

msg.attach(part)
text = msg.as_string()
server = smtplib.SMTP('smtp.gmail.com',587)
server.starttls()
server.login(email_user,email_password)


server.sendmail(email_user,email_send,text)
server.quit()
